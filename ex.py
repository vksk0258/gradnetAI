import io
import os
import PyPDF2
import pikepdf
from google.cloud import vision_v1
from openpyxl import Workbook
# 한 번에 최대 5페이지까지 text 추출 가능

# 첫 페이지는 .txt 파일을 생성해주어야 하므로 1~5 페이지는 따로 처리
def pdf2txt_w(path_num):
    pages = [1,2,3,4,5]
    requests = [{"input_config": input_config, "features": features, "pages": pages}]
    response = client.batch_annotate_files(requests=requests)

    for num, image_response in enumerate(response.responses[0].responses):
#         print(u"Full text: {}".format(image_response.full_text_annotation.text))
        if (num==0):
            with open(destination_path[path_num] + file_name[0:-4] + '.txt', "w",encoding='UTF-8') as f:
                f.write(response.responses[0].responses[num].full_text_annotation.text)
        else:
            with open(destination_path[path_num] + file_name[0:-4] + '.txt', "a",encoding='UTF-8') as f:
                f.write(response.responses[0].responses[num].full_text_annotation.text)

                
# 6페이지 이상인 파일은 아래 함수를 호출하여 텍스트 추출
def pdf2txt(page, path_num):
    pages = [i for i in range(page, page+5)]
    requests = [{"input_config": input_config, "features": features, "pages": pages}]
    response = client.batch_annotate_files(requests=requests)

    for num, image_response in enumerate(response.responses[0].responses):
#         print(u"Full text: {}".format(image_response.full_text_annotation.text))

        with open(destination_path[path_num] + file_name[0:-4] + '.txt', "a",encoding='UTF-8') as f:
                f.write(response.responses[0].responses[num].full_text_annotation.text)

#######################################################################
## source_path와 destination_path에 폴더가 들어가 있으면 에러남
## source_path엔 .pdf 파일만! destination_path는 가급적 빈 폴더!
source_path = ["./input/"]
destination_path = ["./output/"]
#######################################################################

file_count = [0]

for path_num in range(len(source_path)):
    file_list = os.listdir(source_path[path_num])

    file_count.append(file_count[path_num-1]+len(file_list))

    for i in range(len(file_list)):
        file_name = file_list[i]
        client = vision_v1.ImageAnnotatorClient()
        file_path = source_path[path_num] + file_name

        # Supported mime_type: application/pdf, image/tiff, image/gif
        mime_type = "application/pdf"

        try:
            with io.open(file_path, "rb") as f:
                content = f.read()
                pdf_reader = PyPDF2.PdfFileReader(open(file_path, "rb"), strict = False)
        #         pdf_reader = PyPDF2.PdfFileReader(f)
                num_of_pages = pdf_reader.numPages

        except:
            output_path = source_path[path_num] + "decrypted_" + file_list[i]
            file_path = source_path[path_num] + file_name
            pdf = pikepdf.Pdf.new()

            for _, page in enumerate(input_pdf.pages):
                pdf.pages.append(page)

            pdf.save(output_path)
            input_pdf.close()
            print("saved at : {}".format(output_path))

            file_path = source_path[path_num] + "decrypted_" + file_list[i]

            with io.open(file_path, "rb") as f:
                content = f.read()
                pdf_reader = PyPDF2.PdfFileReader(open(file_path, "rb"), strict = False)
        #         pdf_reader = PyPDF2.PdfFileReader(f)
                num_of_pages = pdf_reader.numPages


        input_config = {"mime_type": mime_type, "content": content}
        features = [{"type_": vision_v1.Feature.Type.DOCUMENT_TEXT_DETECTION}]

        for p in range(1, num_of_pages+1, 5):
            if(p==1):
                pdf2txt_w(path_num)
            else:
                pdf2txt(p, path_num)

# 엑셀 파일에 하나로 합치기

txt_source_path = destination_path

#######################################################################
txt_destination_path = "./test/"
label = ["TOEIC", "성적증명서"]
#######################################################################

write_wb = Workbook()
write_ws = write_wb.active

write_ws.cell(1, 2, "파일이름")
write_ws.cell(1, 3, "내용")
write_ws.cell(1, 4, "분류")

for num in range(len(txt_source_path)):
    txt_file_names = os.listdir(txt_source_path[num])
    
    for i, txt_file_name in enumerate(txt_file_names):
        path = txt_source_path[num] + txt_file_name
        r = open(path, 'rt', encoding='UTF-8')
        write_ws.cell(file_count[num] + i + 2, 1, file_count[num] + i + 1)
        write_ws.cell(file_count[num] + i + 2, 2, txt_file_name[:-4])
        write_ws.cell(file_count[num] + i + 2, 3, r.read())
        write_ws.cell(file_count[num] + i + 2, 4, label[num])        
        
#######################################################################
# 엑셀파일로 저장
write_wb.save(txt_destination_path+'output.xlsx')
#######################################################################